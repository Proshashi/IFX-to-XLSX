#!/usr/bin/env python3
"""
ISS .ifx to .xlsx Converter — GUI version for macOS.

Run directly:   python3 ifx_to_xlsx_gui.py
Build .app:     pyinstaller --windowed --name "IFX to Excel" ifx_to_xlsx_gui.py

Output formats (selectable in GUI):
  • MB compatible — single sheet, pivoted matrix (wavelengths × iterations),
    matches the output of MB's ImportFluorimeterSpectra VBA macro.
    Drops IntensityStdError and ExcitationChannel.
  • Comprehensive — three sheets: Spectra (pivoted matrix), Data (full
    long-format with all columns), Metadata (file header).

Intensity values are preserved at full precision in both formats.
"""
import re
import threading
from pathlib import Path
from tkinter import (
    END, E, N, S, W,
    StringVar, Tk, filedialog, messagebox, ttk,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------- Parsing ----------

def parse_ifx(path: Path):
    """Parse one .ifx file. Returns (metadata_dict, columns, rows)."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    if "[Data]" not in text:
        raise ValueError(f"No [Data] section in {path.name}")
    header_text, _, data_text = text.partition("[Data]")

    metadata = {}
    for line in header_text.splitlines():
        s = line.strip()
        if "=" in s:
            k, v = s.split("=", 1)
            metadata[k.strip()] = v.strip()

    columns = [c.strip() for c in metadata.get("Columns", "").split(",") if c.strip()]

    rows = []
    for line in data_text.splitlines():
        s = line.strip()
        if not s:
            continue
        parts = re.split(r"\s+", s)
        parsed = []
        for p in parts:
            try:
                if "." in p or "e" in p.lower():
                    parsed.append(float(p))
                else:
                    parsed.append(int(p))
            except ValueError:
                parsed.append(p)
        rows.append(parsed)
    return metadata, columns, rows


# ---------- Styling helpers ----------

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
BODY_FONT = Font(name="Arial")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row=1, n_cols=None):
    n = n_cols if n_cols is not None else ws.max_column
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


# ---------- Pivot helper ----------

class PivotNotPossible(Exception):
    """Raised when the data layout can't be pivoted to a 2D matrix."""


def _axis_unit(metadata, axis_name):
    """Return unit string (e.g. 'nm') parsed from header line, or ''."""
    raw = metadata.get(axis_name, "")
    m = re.search(r"unit\s*:\s*([^,]+)", raw)
    if m:
        u = m.group(1).strip()
        return "" if u.lower() == "none" else u
    return ""


def _axis_label(axis_name, unit):
    # Turn CamelCase into spaced words, then append unit
    spaced = re.sub(r"(?<!^)(?=[A-Z])", " ", axis_name)
    return f"{spaced} ({unit})" if unit else spaced


def build_pivot(metadata, columns, rows):
    """Return (row_axis_label, col_axis_label, row_vals, col_vals, matrix).

    Driven by the file's own metadata:
      - `Space=` declares the independent axes
      - `Measurable=` names the value column (e.g. Intensity, Anisotropy)

    Rules:
      • Requires exactly 2 axes in `Space` → otherwise raises PivotNotPossible.
      • If "Iteration" is one axis, it becomes the column axis (series).
      • Otherwise the axis with fewer unique values becomes columns.
    """
    space = [s.strip() for s in metadata.get("Space", "").split(",") if s.strip()]
    measurable = metadata.get("Measurable", "").strip()

    # Fallbacks for older/abbreviated files
    if not space:
        # Assume Iteration + first wavelength-like column
        candidates = [c for c in columns
                      if c in ("Iteration", "EmissionWavelength",
                               "ExcitationWavelength", "Time")]
        space = candidates[:2]
    if not measurable:
        for c in columns:
            if c in ("Intensity", "Anisotropy", "Polarization", "Lifetime"):
                measurable = c
                break

    if len(space) != 2:
        raise PivotNotPossible(
            f"Cannot pivot: expected 2 axes in Space, got {len(space)} ({space}). "
            f"Use Comprehensive mode to still get a Data sheet."
        )
    if not measurable or measurable not in columns:
        raise PivotNotPossible(
            f"Cannot pivot: value column {measurable!r} not found in {columns}."
        )
    for ax in space:
        if ax not in columns:
            raise PivotNotPossible(f"Cannot pivot: axis {ax!r} not in columns.")

    # Decide which axis is columns (series) vs rows (primary)
    if "Iteration" in space:
        col_axis = "Iteration"
        row_axis = [a for a in space if a != "Iteration"][0]
    else:
        # Pick fewer-unique as columns
        uniq = {ax: len({r[columns.index(ax)] for r in rows}) for ax in space}
        col_axis = min(uniq, key=uniq.get)
        row_axis = [a for a in space if a != col_axis][0]

    row_idx = columns.index(row_axis)
    col_idx = columns.index(col_axis)
    val_idx = columns.index(measurable)

    row_vals, col_vals = [], []
    row_seen, col_seen = {}, {}
    matrix = {}  # (col_val, row_val) -> value
    for r in rows:
        rv = r[row_idx]
        cv = r[col_idx]
        val = r[val_idx]
        if cv not in col_seen:
            col_seen[cv] = True
            col_vals.append(cv)
        if rv not in row_seen:
            row_seen[rv] = True
            row_vals.append(rv)
        matrix[(cv, rv)] = val

    row_label = _axis_label(row_axis, _axis_unit(metadata, row_axis)) \
                or row_axis
    col_label = _axis_label(col_axis, _axis_unit(metadata, col_axis)) \
                or col_axis

    return row_label, col_label, row_vals, col_vals, matrix, measurable


# ---------- Writers ----------

def _write_spectra_sheet(ws, row_label, col_label, row_vals, col_vals, matrix):
    """Write a pivoted sheet with dynamic axis labels."""
    ws.cell(row=1, column=1, value=row_label)
    for j, cv in enumerate(col_vals, start=2):
        ws.cell(row=1, column=j, value=cv)
    _style_header_row(ws, row=1, n_cols=1 + len(col_vals))

    for i, rv in enumerate(row_vals, start=2):
        ws.cell(row=i, column=1, value=rv)
        for j, cv in enumerate(col_vals, start=2):
            val = matrix.get((cv, rv))
            if val is not None:
                cell = ws.cell(row=i, column=j, value=val)
                if isinstance(val, float):
                    cell.number_format = "0.######"

    for i in range(2, 2 + len(row_vals)):
        c = ws.cell(row=i, column=1)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = max(18, len(row_label) + 2)
    for j in range(2, 2 + len(col_vals)):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = "B2"


def _write_metadata_sheet(ws, metadata):
    ws.append(["Field", "Value"])
    _style_header_row(ws, row=1, n_cols=2)
    for k, v in metadata.items():
        ws.append([k, v])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 70
    ws.freeze_panes = "A2"


def _write_data_sheet(ws, columns, rows):
    if columns:
        ws.append(columns)
    else:
        ncols = max((len(r) for r in rows), default=0)
        ws.append([f"Col{i + 1}" for i in range(ncols)])
    _style_header_row(ws, row=1)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            if isinstance(cell.value, float):
                cell.number_format = "0.######"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        ws.column_dimensions[letter].width = max(14, header_len + 4)
    ws.freeze_panes = "A2"


def _safe_sheet_name(name: str) -> str:
    bad = set(r':\/?*[]')
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31]


def write_mb_format(metadata, columns, rows, out_path: Path):
    """Single sheet, pivoted matrix only. Matches MB's VBA output."""
    row_label, col_label, row_vals, col_vals, matrix, _ = build_pivot(
        metadata, columns, rows
    )
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(out_path.stem)
    _write_spectra_sheet(ws, row_label, col_label, row_vals, col_vals, matrix)
    wb.save(out_path)


def write_comprehensive(metadata, columns, rows, out_path: Path):
    """Three sheets: Spectra (pivot, if possible), Data (long), Metadata.

    If the data can't be pivoted (wrong number of axes, etc.) the Spectra
    sheet is skipped and only Data + Metadata are written.
    """
    wb = Workbook()

    pivot_ok = True
    try:
        row_label, col_label, row_vals, col_vals, matrix, _ = build_pivot(
            metadata, columns, rows
        )
    except PivotNotPossible:
        pivot_ok = False

    if pivot_ok:
        spectra = wb.active
        spectra.title = "Spectra"
        _write_spectra_sheet(spectra, row_label, col_label, row_vals, col_vals, matrix)
        data = wb.create_sheet("Data")
    else:
        data = wb.active
        data.title = "Data"
    _write_data_sheet(data, columns, rows)

    meta = wb.create_sheet("Metadata")
    _write_metadata_sheet(meta, metadata)

    wb.save(out_path)


def convert_file(ifx_path: Path, out_path: Path, mode: str) -> Path:
    """Convert a single .ifx file. out_path is the destination .xlsx."""
    metadata, columns, rows = parse_ifx(ifx_path)
    if mode == "mb":
        try:
            write_mb_format(metadata, columns, rows, out_path)
        except PivotNotPossible as e:
            raise PivotNotPossible(
                f"{e} Try 'Comprehensive' mode instead."
            )
    elif mode == "comprehensive":
        write_comprehensive(metadata, columns, rows, out_path)
    else:
        raise ValueError(f"Unknown mode: {mode}")
    return out_path


# ---------- GUI ----------

class App:
    def __init__(self, root: Tk):
        self.root = root
        root.title("IFX to Excel Converter")
        root.geometry("680x520")
        root.minsize(600, 460)

        self.files = []
        self.out_dir = StringVar(value=str(Path.home() / "Desktop"))
        self.out_name = StringVar(value="")
        self.mode = StringVar(value="mb")

        main = ttk.Frame(root, padding=12)
        main.grid(row=0, column=0, sticky=(N, S, E, W))
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(3, weight=1)

        ttk.Label(
            main, text="ISS .ifx → .xlsx Converter",
            font=("Helvetica", 16, "bold"),
        ).grid(row=0, column=0, sticky=W, pady=(0, 4))
        ttk.Label(
            main,
            text="Select one or more .ifx files, choose an output format, and convert.",
            foreground="#555",
        ).grid(row=1, column=0, sticky=W, pady=(0, 10))

        file_row = ttk.Frame(main)
        file_row.grid(row=2, column=0, sticky=(E, W), pady=4)
        file_row.columnconfigure(1, weight=1)
        ttk.Button(
            file_row, text="Choose .ifx files…", command=self.choose_files
        ).grid(row=0, column=0, sticky=W)
        ttk.Button(
            file_row, text="Clear", command=self.clear_files
        ).grid(row=0, column=2, sticky=E, padx=(6, 0))
        self.count_label = ttk.Label(file_row, text="No files selected")
        self.count_label.grid(row=0, column=1, sticky=W, padx=10)

        list_frame = ttk.Frame(main)
        list_frame.grid(row=3, column=0, sticky=(N, S, E, W), pady=6)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        from tkinter import Listbox
        self.listbox = Listbox(list_frame, activestyle="none", highlightthickness=0)
        self.listbox.grid(row=0, column=0, sticky=(N, S, E, W))
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        sb.grid(row=0, column=1, sticky=(N, S))
        self.listbox.config(yscrollcommand=sb.set)

        fmt_frame = ttk.LabelFrame(main, text="Output format", padding=8)
        fmt_frame.grid(row=4, column=0, sticky=(E, W), pady=(8, 4))
        ttk.Radiobutton(
            fmt_frame,
            text="MB compatible — pivoted matrix only "
                 "(wavelengths × iterations, no metadata)",
            variable=self.mode,
            value="mb",
        ).grid(row=0, column=0, sticky=W)
        ttk.Radiobutton(
            fmt_frame,
            text="Comprehensive — Spectra (pivot) + Data (long) + Metadata",
            variable=self.mode,
            value="comprehensive",
        ).grid(row=1, column=0, sticky=W)

        out_row = ttk.Frame(main)
        out_row.grid(row=5, column=0, sticky=(E, W), pady=4)
        out_row.columnconfigure(1, weight=1)
        ttk.Label(out_row, text="Save to:").grid(row=0, column=0, sticky=W)
        ttk.Entry(out_row, textvariable=self.out_dir).grid(
            row=0, column=1, sticky=(E, W), padx=6
        )
        ttk.Button(
            out_row, text="Browse…", command=self.choose_out_dir
        ).grid(row=0, column=2)

        self.name_row = ttk.Frame(main)
        self.name_row.grid(row=6, column=0, sticky=(E, W), pady=4)
        self.name_row.columnconfigure(1, weight=1)
        ttk.Label(self.name_row, text="Output name:").grid(row=0, column=0, sticky=W)
        self.name_entry = ttk.Entry(self.name_row, textvariable=self.out_name)
        self.name_entry.grid(row=0, column=1, sticky=(E, W), padx=6)
        ttk.Label(self.name_row, text=".xlsx", foreground="#888").grid(
            row=0, column=2, sticky=W
        )
        self.name_row.grid_remove()

        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.grid(row=7, column=0, sticky=(E, W), pady=(10, 4))
        self.status = StringVar(value="Ready.")
        ttk.Label(main, textvariable=self.status, foreground="#444").grid(
            row=8, column=0, sticky=W
        )

        self.convert_btn = ttk.Button(
            main, text="Convert", command=self.start_convert
        )
        self.convert_btn.grid(row=9, column=0, sticky=E, pady=(10, 0))

    def choose_files(self):
        paths = filedialog.askopenfilenames(
            title="Select .ifx files",
            filetypes=[("IFX files", "*.ifx"), ("All files", "*.*")],
        )
        if paths:
            self.files = [Path(p) for p in paths]
            self.refresh_list()

    def clear_files(self):
        self.files = []
        self.refresh_list()

    def refresh_list(self):
        self.listbox.delete(0, END)
        for p in self.files:
            self.listbox.insert(END, p.name)
        n = len(self.files)
        self.count_label.config(
            text=f"{n} file{'s' if n != 1 else ''} selected" if n else "No files selected"
        )
        if n == 1:
            self.out_name.set(self.files[0].stem)
            self.name_row.grid()
        else:
            self.out_name.set("")
            self.name_row.grid_remove()

    def choose_out_dir(self):
        d = filedialog.askdirectory(
            initialdir=self.out_dir.get(), title="Choose output folder"
        )
        if d:
            self.out_dir.set(d)

    def _on_ui(self, fn, *args, **kwargs):
        """Run fn on the Tk main thread; block worker until it returns."""
        result = {}
        done = threading.Event()

        def run():
            try:
                result["value"] = fn(*args, **kwargs)
            except Exception as e:
                result["exc"] = e
            finally:
                done.set()

        self.root.after(0, run)
        done.wait()
        if "exc" in result:
            raise result["exc"]
        return result.get("value")

    def _resolve_out_path(self, ifx_path: Path, out_dir: Path) -> Path:
        """Pick the destination filename — custom name only if 1 file, else source stem."""
        custom = self.out_name.get().strip() if len(self.files) == 1 else ""
        stem = custom or ifx_path.stem
        if stem.lower().endswith(".xlsx"):
            stem = stem[:-5]
        return out_dir / f"{stem}.xlsx"

    def start_convert(self):
        if not self.files:
            messagebox.showwarning("No files", "Please select at least one .ifx file.")
            return
        out_dir = Path(self.out_dir.get()).expanduser()
        if not out_dir.exists():
            try:
                out_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"Cannot create output folder:\n{e}")
                return
        self.convert_btn.config(state="disabled")
        self.progress.config(maximum=len(self.files), value=0)
        threading.Thread(
            target=self._run_conversion, args=(out_dir, self.mode.get()), daemon=True
        ).start()

    def _run_conversion(self, out_dir: Path, mode: str):
        done, failed, skipped = [], [], []
        for i, f in enumerate(self.files, 1):
            out_path = self._resolve_out_path(f, out_dir)
            if out_path.exists():
                proceed = self._on_ui(
                    messagebox.askyesno,
                    "File exists",
                    f"{out_path.name} already exists in:\n{out_dir}\n\nOverwrite?",
                )
                if not proceed:
                    skipped.append(out_path.name)
                    self.progress.config(value=i)
                    self.root.update_idletasks()
                    continue
            self.status.set(f"Converting {f.name} ({i}/{len(self.files)})…")
            try:
                out = convert_file(f, out_path, mode)
                done.append(out)
            except Exception as e:
                failed.append((f.name, str(e)))
            self.progress.config(value=i)
            self.root.update_idletasks()

        self.convert_btn.config(state="normal")
        parts = [f"Converted {len(done)} file(s)."]
        if skipped:
            parts.append("\nSkipped (already existed):\n" + "\n".join(f"• {n}" for n in skipped))
        if failed:
            parts.append("\nFailed:\n" + "\n".join(f"• {n}: {err}" for n, err in failed))
        msg = "".join(parts)
        if failed:
            self.status.set(
                f"Done with errors — {len(done)} ok, {len(skipped)} skipped, {len(failed)} failed."
            )
            self._on_ui(messagebox.showwarning, "Completed with errors", msg)
        elif skipped:
            self.status.set(f"Done — {len(done)} ok, {len(skipped)} skipped.")
            self._on_ui(
                messagebox.showinfo,
                "Completed",
                msg + f"\n\nSaved to:\n{out_dir}",
            )
        else:
            self.status.set(f"Done. {len(done)} file(s) saved to {out_dir}")
            self._on_ui(
                messagebox.showinfo,
                "Success",
                f"Converted {len(done)} file(s).\n\nSaved to:\n{out_dir}",
            )


def main():
    root = Tk()
    try:
        style = ttk.Style()
        if "aqua" in style.theme_names():
            style.theme_use("aqua")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
